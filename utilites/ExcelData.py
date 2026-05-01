import openpyxl
class ExcelData:


    def getRowCount(self,filename,sheetname):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]
        return sheet.max_row

    def getColumnCount(self,filename,sheetname):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]
        return sheet.max_column

    def getCellData(self,filename, sheetname, rowNumber, columnNumber):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]
        return sheet.cell(row=rowNumber, column=columnNumber).value

    def setCellData(self,filename, sheetname, rowNumber, columnNumber,inputData):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]
        sheet.cell(row=rowNumber, column=columnNumber).value = inputData
        workbook.save(filename)

# print(getRowCount("Book1.xlsx","Sheet1"))



